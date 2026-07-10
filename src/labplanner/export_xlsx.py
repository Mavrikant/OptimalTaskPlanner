"""Build a styled .xlsx workbook from tabular data (used by the schedule export)."""

from __future__ import annotations

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MAX_ROWS = 20_000
MAX_COLS = 64
_INVALID_SHEET = re.compile(r"[\\/*?:\[\]]")


def _sheet_title(name: str) -> str:
    title = _INVALID_SHEET.sub(" ", name or "").strip()
    return title[:31] or "Sheet1"


def build_workbook(
    columns: list[str],
    rows: list[list],
    sheet_name: str = "Sheet1",
    chart_png: bytes | None = None,
    chart_sheet_name: str = "Chart",
    chart_size: tuple[int, int] | None = None,
) -> bytes:
    """Return the bytes of an .xlsx workbook.

    With a chart image, sheet 1 holds the Gantt picture and sheet 2 the table;
    otherwise a single table sheet. The table header is bold, frozen and
    auto-filtered, and columns are auto-sized.
    """
    columns = [str(c) for c in columns[:MAX_COLS]]
    ncols = len(columns)
    wb = Workbook()

    if chart_png:
        ws_chart = wb.active
        ws_chart.title = _sheet_title(chart_sheet_name)
        img = XlImage(BytesIO(chart_png))
        if chart_size:
            img.width, img.height = chart_size
        ws_chart.add_image(img, "A1")
        ws = wb.create_sheet(_sheet_title(sheet_name))
    else:
        ws = wb.active
        ws.title = _sheet_title(sheet_name)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    ws.append(columns)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    widths = [len(c) for c in columns]
    for row in rows[:MAX_ROWS]:
        values = [_coerce(v) for v in row[:ncols]]
        values += [None] * (ncols - len(values))
        ws.append(values)
        for i, v in enumerate(values):
            if v is not None:
                widths[i] = max(widths[i], len(str(v)))

    if ncols:
        last_col = get_column_letter(ncols)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(60, max(10, width + 2))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _coerce(value):
    """Keep numbers numeric so Excel treats them as numbers; everything else -> str."""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) or value is None:
        return value
    return str(value)
