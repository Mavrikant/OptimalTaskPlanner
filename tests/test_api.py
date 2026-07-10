from labplanner import __version__


def default_pid(client):
    return client.get("/api/projects").json()["projects"][0]["id"]


def test_health(client):
    r = client.get("/api/health")
    assert r.status_code == 200
    assert r.json() == {"status": "ok", "version": __version__}


def test_index_serves_frontend(client):
    r = client.get("/")
    assert r.status_code == 200
    assert "LabPlanner" in r.text


def test_default_project_is_seeded(client):
    projects = client.get("/api/projects").json()["projects"]
    assert len(projects) == 1
    d = client.get(f"/api/projects/{projects[0]['id']}").json()
    assert d["project"]["equipment"]
    assert d["horizon"]["days"] == 14
    assert d["horizon"]["slots_per_day"] == 48
    assert len(d["horizon"]["day_dates"]) == 14


def test_project_crud_flow(client):
    r = client.post("/api/projects", json={"name": "Bench B"})
    assert r.status_code == 200
    pid = r.json()["id"]

    assert client.patch(f"/api/projects/{pid}", json={"name": "Bench B2"}).status_code == 200
    names = {p["id"]: p["name"] for p in client.get("/api/projects").json()["projects"]}
    assert names[pid] == "Bench B2"

    dup = client.post(f"/api/projects/{pid}/duplicate").json()
    assert dup["name"] == "Bench B2 (copy)"

    assert client.delete(f"/api/projects/{pid}").status_code == 200
    assert client.get(f"/api/projects/{pid}").status_code == 404


def test_unknown_project_is_404(client):
    assert client.get("/api/projects/nope").status_code == 404
    assert client.put("/api/projects/nope", json={"name": "x"}).status_code == 404
    assert client.post("/api/projects/nope/solve").status_code == 404
    assert client.delete("/api/projects/nope").status_code == 404


def test_put_project_roundtrip(client):
    pid = default_pid(client)
    project = client.get(f"/api/projects/{pid}").json()["project"]
    project["calendar"]["work_start"] = "09:30"
    project["calendar"]["holidays"] = ["2026-12-25"]

    r = client.put(f"/api/projects/{pid}", json=project)
    assert r.status_code == 200
    assert r.json()["horizon"]["work_start_slot"] == 19

    again = client.get(f"/api/projects/{pid}").json()["project"]
    assert again["calendar"]["work_start"] == "09:30"
    assert again["calendar"]["holidays"] == ["2026-12-25"]


def test_put_rejects_invalid_project(client):
    pid = default_pid(client)
    project = client.get(f"/api/projects/{pid}").json()["project"]
    project["tasks"][0]["minutes"] = 45  # not slot-aligned
    assert client.put(f"/api/projects/{pid}", json=project).status_code == 422


def test_import_project(client):
    pid = default_pid(client)
    data = client.get(f"/api/projects/{pid}").json()["project"]
    data["name"] = "Imported copy"
    r = client.post("/api/projects/import", json=data)
    assert r.status_code == 200
    assert r.json()["name"] == "Imported copy"

    assert client.post("/api/projects/import",
                       json={"tasks": [{"id": "x", "name": "bad", "minutes": 45}]}
                       ).status_code == 422


def test_solve_returns_and_persists_schedule(client):
    pid = default_pid(client)
    r = client.post(f"/api/projects/{pid}/solve")
    assert r.status_code == 200
    schedule = r.json()["schedule"]
    assert schedule["status"] in ("OPTIMAL", "FEASIBLE")
    assert schedule["tasks"]

    d = client.get(f"/api/projects/{pid}").json()
    assert d["project"]["schedule"]["status"] == schedule["status"]


def test_backups_listed_and_restored(client):
    pid = default_pid(client)
    project = client.get(f"/api/projects/{pid}").json()["project"]
    original = project["calendar"]["work_start"]

    # deleting-like force snapshot: rename triggers save -> time-based snapshot
    client.patch(f"/api/projects/{pid}", json={"name": "Snap"})
    backups = client.get(f"/api/projects/{pid}/backups").json()["backups"]
    assert backups, "a snapshot should exist after the first save"

    project["calendar"]["work_start"] = "10:30"
    project["name"] = "Snap"
    client.put(f"/api/projects/{pid}", json=project)

    r = client.post(f"/api/projects/{pid}/backups/{backups[0]['name']}/restore")
    assert r.status_code == 200
    restored = client.get(f"/api/projects/{pid}").json()["project"]
    assert restored["calendar"]["work_start"] == original

    assert client.post(f"/api/projects/{pid}/backups/evil.json/restore").status_code == 404


def test_export_xlsx_roundtrips(client):
    from io import BytesIO

    from openpyxl import load_workbook

    payload = {
        "filename": "sched",
        "sheet_name": "Schedule",
        "columns": ["Task", "Start", "Duration"],
        "rows": [["RF sweep", "2026-07-13 08:00", 4], ["Soak", "2026-07-14 09:00", 36]],
    }
    r = client.post("/api/export/xlsx", json=payload)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert r.headers["content-disposition"].endswith('filename="sched.xlsx"')

    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    assert ws.title == "Schedule"
    assert [c.value for c in ws[1]] == ["Task", "Start", "Duration"]
    assert ws["A2"].value == "RF sweep"
    assert ws["C2"].value == 4  # numbers stay numeric
    assert ws.freeze_panes == "A2"


def test_export_xlsx_with_chart_has_two_sheets(client):
    from io import BytesIO

    from openpyxl import load_workbook

    png = ("data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwC"
           "AAAAC0lEQVR42mNk+M9QDwADhgGAWjR9awAAAABJRU5ErkJggg==")
    payload = {
        "filename": "sched.xlsx", "sheet_name": "Tasks", "chart_sheet_name": "Chart",
        "chart_png_base64": png, "chart_w": 400, "chart_h": 200,
        "columns": ["Task"], "rows": [["RF sweep"]],
    }
    r = client.post("/api/export/xlsx", json=payload)
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    assert wb.sheetnames == ["Chart", "Tasks"]
    assert len(wb["Chart"]._images) == 1
    assert wb["Tasks"]["A1"].value == "Task"


def test_holiday_countries(client):
    d = client.get("/api/holidays/countries").json()
    codes = {c["code"] for c in d["countries"]}
    assert "TR" in codes
    assert "US" in codes


def test_holidays_for_turkey(client):
    d = client.get("/api/holidays", params={"country": "TR", "year": 2026}).json()
    dates = {h["date"] for h in d["holidays"]}
    assert "2026-01-01" in dates
    assert "2026-10-29" in dates  # Republic Day


def test_holidays_unknown_country(client):
    r = client.get("/api/holidays", params={"country": "XX", "year": 2026})
    assert r.status_code == 400
